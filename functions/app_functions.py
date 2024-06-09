# PEP8 Compliant Guidance
# Standard Library Imports
import os
import logging
import tkinter as tk
import datetime as dt
from tkinter import messagebox

# Third-Party Library Imports
import openpyxl
import customtkinter as ctk # type: ignore
import win32com.client as win32 # type: ignore

# Local Application/Library Specific Imports
from constants import log_file
from constants import COLOR_SPECS

# Logging Format
logging.basicConfig(level=logging.ERROR, 
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    filename=log_file, 
                    filemode='a'
)

# Set up logging to print to the console
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def lock_widgets(container):
    """
    Lock (disable) all Button and Entry widgets in the given container.

    Args:
        container (tk.BaseWidget): The container widget.
    """
    for child in container.winfo_children():
        if isinstance(child, (ctk.CTkButton, tk.Button)):
            child.configure(state=tk.DISABLED)
        elif isinstance(child, (tk.Entry, ctk.CTkEntry)):
            child.configure(state=tk.DISABLED)
        lock_widgets(child)

def lock_and_color_entry_widgets(container):
    """
    Lock (disable) all Entry widgets and apply formatting in the given container and its child frames.

    Args:
        container (tk.BaseWidget): The container widget.
    """
    for child in container.winfo_children():
        if isinstance(child, (tk.Entry, ctk.CTkEntry)):
            entry_text = child.get()
            if entry_text in COLOR_SPECS:
                bg_color = COLOR_SPECS[entry_text]["label_bg"]
                fg_color = COLOR_SPECS[entry_text]["label_text"]
            else:
                bg_color = "white"
                fg_color = "black"
            child.configure(state=tk.DISABLED, disabledbackground=bg_color, disabledforeground=fg_color)
        elif isinstance(child, tk.Frame):
            lock_and_color_entry_widgets(child)  # Recursively call the function for child frames

def get_workbook_info(crew, month, year, schedule):
    """
    Get the workbook filename and worksheet name based on the user selections.

    Args:
        crew (str): The selected crew.
        month (int or datetime): The selected month as an integer or a datetime object.
        year (int or datetime): The selected year as an integer or a datetime object.
        schedule (str): The selected schedule type ("Overtime" or "work_schedule").

    Returns:
        tuple[str, str]: A tuple containing the workbook filename and worksheet name.
    """
    
    if isinstance(month, dt.datetime):
        month_value = month.month
    else:
        month_value = month
    
    worksheet_name = f"{month_value}"

    return worksheet_name

def generate_row_data(schedule, month):
    """
    Generate the row data for the specified schedule and month.

    Args:
        schedule (str): The schedule type ("Overtime" or "work_schedule").
        month (int): The month number.

    Returns:
        list: The row data for the worksheet.
    """
    if schedule == "Overtime":
        row_data = ["Name", "Starting Asking Hours", "Starting Working Hours"] + [f"Day {i+1}" for i in range(31)] + ["Total Asking Hours", "Total Working Hours"]
    elif schedule == "work_schedule":
        row_data = ["Name"] + [f"Day {i+1}" for i in range(31)]
    else:
        row_data = []  # Handle invalid schedule type

    return row_data

def create_workbook(workbook_filename, schedule):
    """
    Create a new workbook with the specified filename and populate it with worksheets based on the schedule type.

    Args:
        workbook_filename (str): The filename for the new workbook.
        schedule (str): The schedule type ("Overtime" or "work_schedule").
    """
    try:
        workbook = openpyxl.Workbook()

        # Remove the default sheet created by openpyxl
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for month in range(1, 13):
            sheet_name = f"Month {month}"
            sheet = workbook.create_sheet(title=sheet_name)
            
            # generate_row_data to generate sheet content
            row_data = generate_row_data(schedule, month)
            sheet.append(row_data)

        workbook.save(workbook_filename)
    except Exception as e:
        logging.error(f"Failed to create workbook '{workbook_filename}': {str(e)}")        

def forward_outlook_email(subject, body, recipient):
    """
    Forward an Outlook email with the specified subject, body, and recipient.

    Args:
        subject (str): The subject of the email.
        body (str): The body content of the email.
        recipient (str): The email address of the recipient.
    """
    try:
        # Create an Outlook application instance
        outlook = win32.Dispatch('outlook.application')

        # Create a new mail item
        mail = outlook.CreateItem(0)

        # Set the subject, body, and recipient of the email
        mail.Subject = subject
        mail.Body = body
        mail.To = recipient
        
        # Display the email window for review
        mail.Display()

        print("Email forwarded successfully.")
    except Exception as e:
        logging.error(f"An error occurred while forwarding the email: {str(e)}")

def assignment_code_formatting(frame, label, code_text, color_specs):
    """
    Format the WSMatrix Legend labels with the appropriate formatting.

    Args:
        frame (widget): The passed frame.
        label (widget): The passed label.
        code_text (str): The text contained in the label.
        color_specs (dict): color specifications {code_text:{bg:color, text:color}}
    """
    if code_text in color_specs:
        frame.configure(fg_color=color_specs[code_text]["frame"])
        label.configure(fg_color=color_specs[code_text]["label_bg"], text_color=color_specs[code_text]["label_text"])
    else:
        frame.configure(fg_color="blue")
        label.configure(fg_color="blue", text_color="white")
        
def apply_entry_color_specs(entry, entry_text):
    """
    Format the WSMatrix Entries with the appropriate formatting.

    Args:
        entry_text (str): The string passed to the function from the entry
        color_specs (dict): color specifications {code_text:{bg:color, text:color}}
    """
    if entry_text in COLOR_SPECS:
        entry_bg = COLOR_SPECS[entry_text]["label_bg"]
        entry_text_color = COLOR_SPECS[entry_text]["label_text"]
        entry.configure(bg=entry_bg, fg=entry_text_color)
    else:
        entry.configure(bg="white", fg="black")  # Default formatting

#TODO: Set Customized themes
def set_theme(root, color, widget_type=None):
    """
    Set the background color of all widgets in the given container.

    Args:
        root (tk.BaseWidget): The container widget.
        color (str): The background color to be set.
        widget_type (type | None, optional): The specific type of widgets to apply the theme to. If None, all widgets will be affected.
    """
    def apply_theme(widget):
        if widget_type is None or isinstance(widget, widget_type):
            try:
                widget.config(bg=color)
            except tk.TclError:
                pass  # Ignore widgets that don't support the bg option
        for child in widget.winfo_children():
            apply_theme(child)

    apply_theme(root)
 
def read_save_data(data_file, ws_name):
    """
    Read the number of names from an Excel file.

    Args:
        data_file (str): The file path of the Excel file.
        ws_name (str): The name of the worksheet.

    Returns:
        int | None: The number of names in the Excel file, or None if an error occurred.
    """
    try:
        if not os.path.exists(data_file):
            print("New workbook message at read_save_data functions.py")
            
        return None
    
    except PermissionError:
        messagebox.showerror("File in Use", f"The file '{data_file}' is currently in use. Please close the file and try again.")
        logging.error(f"File Appears to be open. Close to read '{data_file}'")

    except FileNotFoundError:
        logging.error(f"File '{data_file}' not found.")
        messagebox.showerror("File Not Found", f"{data_file} not found.\nPlease view the log or reload the schedule.")
    except KeyError:
        logging.error(f"Worksheet '{ws_name}' not found in the workbook.")
        messagebox.showerror("Worksheet Not Found", f"Worksheet '{ws_name}' not found in the workbook.\nPlease check the worksheet name.")

    except Exception as e:
        logging.error(f"Failed to read data from file '{data_file}': {str(e)}")
        messagebox.showerror("Unknown Error", "An error occurred.\nPlease see log for details.")
    return None
        
def get_hours_for_selected_person(workbook_filename, worksheet_name, selected_name):
    try:
        workbook = openpyxl.load_workbook(workbook_filename)
        worksheet = workbook[worksheet_name]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] == selected_name:
                starting_asking_hours = row[1]
                starting_working_hours = row[2]
                return starting_working_hours, starting_asking_hours

        return None, None  # Return None if the selected name is not found
    
    except PermissionError:
        messagebox.showerror("File in Use", f"The file '{workbook_filename}' is currently in use. Please close the file and try again.")
        logging.error(f"PermissionError opening or loading Overtime Schedule '{workbook_filename}'.")
    except FileNotFoundError:
        messagebox.showerror("Error", f"Workbook '{workbook_filename}' not found.")
        logging.error(f"File '{workbook_filename} not found'.")
        return None, None
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
        logging.error(f"Error opening or loading Overtime Schedule '{workbook_filename}'.\nException: {str(e)}")
        return None, None
    finally:
        if 'workbook' in locals():
            workbook.close()

def center_toplevel_window(toplevel):
    """
    Center a Toplevel window on the screen.

    Args:
        toplevel (tk.Toplevel): The Toplevel window to be centered.
    """
    # Update the window's dimensions
    toplevel.update_idletasks()

    # Get the window's width and height
    window_width = toplevel.winfo_width()
    window_height = toplevel.winfo_height()

    # Get the screen's width and height
    screen_width = toplevel.winfo_screenwidth()
    screen_height = toplevel.winfo_screenheight()

    # Calculate the position to center the window
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)

    # Set the window's position
    toplevel.geometry(f"+{position_x}+{position_y}")

""" 
if the name of the file is main.py, then run the code
in this case, it won't run because the file is functions.py 
"""
# Test Code
if __name__ == "__main__":
    pass
